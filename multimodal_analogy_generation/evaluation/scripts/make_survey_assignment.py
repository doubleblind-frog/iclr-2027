"""
make_survey_assignment.py

Generates a balanced participant assignment for a Qualtrics image-pair
rating survey.

Usage:
    python make_survey_assignment.py
    python multimodal_analogy_generation/evaluation/scripts/make_survey_assignment.py --zero   multimodal_analogy_generation/evaluation/results/zero_filtered --last   multimodal_analogy_generation/evaluation/results/last_filtered --lookup multimodal_analogy_generation/zero_shot_exploration/zero_shot_outputs.jsonl --out    multimodal_analogy_generation/evaluation/results/survey_assignment --seed   42
"""

from __future__ import annotations

import argparse
import csv
import json
import random
from collections import Counter
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

ZERO_DIR    = Path("multimodal_analogy_generation/evaluation/results/zero_filtered")
LAST_DIR    = Path("multimodal_analogy_generation/evaluation/results/last_filtered")
LOOKUP_PATH = Path("multimodal_analogy_generation/zero_shot_exploration/zero_shot_outputs.jsonl")
OUT_STEM    = Path("multimodal_analogy_generation/evaluation/results/survey_assignment")
N_PER_GROUP = 5
N_REVIEWS   = 2


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_metaphor_lookup(path: Path) -> dict[str, str]:
    """Map metaphor id -> metaphor text from zero_shot_outputs.jsonl."""
    lookup: dict[str, str] = {}
    if not path.exists():
        print(f"  Warning: lookup not found at {path} — metaphor text will be blank.")
        return lookup
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            row  = json.loads(line)
            mid  = str(row.get("id", "")).strip()
            text = (row.get("metaphor") or "").strip()
            if mid:
                lookup[mid] = text
    return lookup


# ---------------------------------------------------------------------------
# Core assignment
# ---------------------------------------------------------------------------

def make_assignment(ids: list[str], seed: int) -> list[list[str]]:
    rng = random.Random(seed)
    n   = len(ids)
    assert n % N_PER_GROUP == 0
    n_groups_per_pass = n // N_PER_GROUP
    all_groups: list[list[str]] = []
    for _ in range(N_REVIEWS):
        shuffled = ids.copy()
        rng.shuffle(shuffled)
        for i in range(n_groups_per_pass):
            all_groups.append(shuffled[i * N_PER_GROUP : (i + 1) * N_PER_GROUP])
    rng.shuffle(all_groups)
    return all_groups


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------

def verify(groups: list[list[str]], ids: list[str]) -> None:
    all_ids = set(ids)
    for i, g in enumerate(groups, 1):
        assert len(g) == N_PER_GROUP
        assert len(set(g)) == len(g), f"Group {i} has duplicate pairs: {g}"
        for mid in g:
            assert mid in all_ids, f"Group {i}: unknown ID {mid!r}"
    counts = Counter(mid for g in groups for mid in g)
    for mid in ids:
        assert counts[mid] == N_REVIEWS, \
            f"ID {mid!r} appears {counts[mid]}×, expected {N_REVIEWS}"
    print(f"  ✓ {len(groups)} groups · {N_PER_GROUP} pairs each · "
          f"{len(ids)} unique pairs · each reviewed {N_REVIEWS}× — balanced.")


# ---------------------------------------------------------------------------
# Outputs
# ---------------------------------------------------------------------------

def write_csv(groups: list[list[str]], path: Path,
              zero_dir: Path, last_dir: Path,
              lookup: dict[str, str]) -> None:
    header = (
        ["group_id"]
        + [f"pair_{n}_id"       for n in range(1, N_PER_GROUP + 1)]
        + [f"pair_{n}_metaphor" for n in range(1, N_PER_GROUP + 1)]
        + [f"pair_{n}_zero_image" for n in range(1, N_PER_GROUP + 1)]
        + [f"pair_{n}_last_image" for n in range(1, N_PER_GROUP + 1)]
    )
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for i, group in enumerate(groups, 1):
            w.writerow(
                [i]
                + group
                + [lookup.get(mid, "") for mid in group]
                + [str(zero_dir / f"{mid}.png") for mid in group]
                + [str(last_dir  / f"{mid}.png") for mid in group]
            )
    print(f"  CSV  → {path}")


def write_xlsx(groups: list[list[str]], path: Path,
               zero_dir: Path, last_dir: Path,
               lookup: dict[str, str]) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  openpyxl not installed — skipping Excel output.")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participant Assignment"
    header = (
        ["group_id"]
        + [f"pair_{n}_id"         for n in range(1, N_PER_GROUP + 1)]
        + [f"pair_{n}_metaphor"   for n in range(1, N_PER_GROUP + 1)]
        + [f"pair_{n}_zero_image" for n in range(1, N_PER_GROUP + 1)]
        + [f"pair_{n}_last_image" for n in range(1, N_PER_GROUP + 1)]
    )
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    alt_fill = PatternFill("solid", fgColor="DCE6F1")
    for i, group in enumerate(groups, 1):
        row_data = (
            [i]
            + group
            + [lookup.get(mid, "") for mid in group]
            + [str(zero_dir / f"{mid}.png") for mid in group]
            + [str(last_dir  / f"{mid}.png") for mid in group]
        )
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="left")
    ws.column_dimensions["A"].width = 10
    for col in range(2, N_PER_GROUP + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    for col in range(N_PER_GROUP + 2, N_PER_GROUP * 2 + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 52
    for col in range(N_PER_GROUP * 2 + 2, len(header) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 48
    ws.freeze_panes = "B2"
    wb.save(path)
    print(f"  XLSX → {path}")


def write_js(groups: list[list[str]], path: Path,
             lookup: dict[str, str]) -> None:
    """
    Qualtrics JavaScript that randomly assigns a participant to a group
    and writes all embedded data fields.
    """
    lines = [
        "Qualtrics.SurveyEngine.addOnReady(function () {",
        "",
        "    // Assignment lookup: group_id -> [pair_1_id, ..., pair_5_id]",
        "    var ids = {",
    ]
    for i, group in enumerate(groups, 1):
        comma = "," if i < len(groups) else ""
        pair_str = ", ".join(f'"{mid}"' for mid in group)
        lines.append(f"        {i}: [{pair_str}]{comma}")
    lines += [
        "    };",
        "",
        "    // Metaphor text lookup: group_id -> [pair_1_metaphor, ..., pair_5_metaphor]",
        "    var metaphors = {",
    ]
    for i, group in enumerate(groups, 1):
        comma = "," if i < len(groups) else ""
        texts = [
            lookup.get(mid, "").replace("\\", "\\\\").replace('"', '\\"')
            for mid in group
        ]
        text_str = ", ".join(f'"{t}"' for t in texts)
        lines.append(f"        {i}: [{text_str}]{comma}")
    lines += [
        "    };",
        "",
        f"    // Pick a random group (1\u2013{len(groups)})",
        f"    var groupId = Math.ceil(Math.random() * {len(groups)});",
        "    var pairIds  = ids[groupId];",
        "    var pairMets = metaphors[groupId];",
        "",
        "    // Write all embedded data fields",
        "    Qualtrics.SurveyEngine.setEmbeddedData('group_id', String(groupId));",
    ]
    for n in range(1, N_PER_GROUP + 1):
        lines.append(
            f"    Qualtrics.SurveyEngine.setEmbeddedData("
            f"'pair_{n}_id',       pairIds[{n-1}]);"
        )
        lines.append(
            f"    Qualtrics.SurveyEngine.setEmbeddedData("
            f"'pair_{n}_metaphor', pairMets[{n-1}]);"
        )
    lines += ["", "});"]
    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  JS   → {path}")


def write_coverage_report(groups: list[list[str]],
                          ids: list[str], path: Path) -> None:
    coverage: dict[str, list[int]] = {mid: [] for mid in ids}
    for i, group in enumerate(groups, 1):
        for mid in group:
            coverage[mid].append(i)
    path.write_text(json.dumps(coverage, indent=2, ensure_ascii=False),
                    encoding="utf-8")
    print(f"  Coverage → {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--zero",   type=Path, default=ZERO_DIR)
    parser.add_argument("--last",   type=Path, default=LAST_DIR)
    parser.add_argument("--lookup", type=Path, default=LOOKUP_PATH,
                        help="JSONL mapping id -> metaphor text "
                             f"(default: {LOOKUP_PATH})")
    parser.add_argument("--out",    type=Path, default=OUT_STEM,
                        help="Output path stem — extensions added automatically.")
    parser.add_argument("--seed",   type=int,  default=42)
    args = parser.parse_args()

    zero_ids = {p.stem for p in args.zero.glob("*.png")}
    last_ids  = {p.stem for p in args.last.glob("*.png")}
    ids = sorted(zero_ids & last_ids)

    print(f"Zero-shot images : {len(zero_ids)}")
    print(f"Last-round images: {len(last_ids)}")
    if zero_ids - last_ids:
        print(f"  Warning: {len(zero_ids - last_ids)} IDs only in zero — excluded.")
    if last_ids - zero_ids:
        print(f"  Warning: {len(last_ids - zero_ids)} IDs only in last — excluded.")
    print(f"Matched pairs    : {len(ids)}")

    if not ids:
        print("No matched pairs. Check --zero and --last paths.")
        return

    if (len(ids) * N_REVIEWS) % N_PER_GROUP != 0:
        print(f"Error: {len(ids)} × {N_REVIEWS} is not divisible by {N_PER_GROUP}.")
        return

    lookup   = load_metaphor_lookup(args.lookup)
    missing  = [mid for mid in ids if mid not in lookup]
    if missing:
        print(f"  Warning: {len(missing)} IDs have no metaphor text: {missing[:5]}…")

    n_groups = (len(ids) * N_REVIEWS) // N_PER_GROUP
    print(f"Assignment: {len(ids)} pairs × {N_REVIEWS} reviews "
          f"÷ {N_PER_GROUP} per participant = {n_groups} participants\n")

    groups = make_assignment(ids, seed=args.seed)
    verify(groups, ids)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    print()
    write_csv(groups,  args.out.with_suffix(".csv"),  args.zero, args.last, lookup)
    write_xlsx(groups, args.out.with_suffix(".xlsx"), args.zero, args.last, lookup)
    write_js(groups,   args.out.with_name(args.out.stem + "_qualtrics.js"), lookup)
    write_coverage_report(groups, ids, args.out.parent / "survey_coverage.json")
    print(f"\nDone. Recruit {n_groups} participants.")


if __name__ == "__main__":
    main()